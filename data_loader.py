"""Lectura del archivo esquema.xlsx."""
import openpyxl

from config import EXCEL_PATH


def load_excel_data():
    """Carga las cuatro hojas del Excel y las devuelve como listas de dicts.

    Returns:
        (notas, datos, profesores, plan)
    """
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

    # Notas
    ws = wb["notas"]
    notas = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None and row[1] is None:
            continue
        notas.append({
            "profesor": row[0] or "",
            "alumna": row[1] or "",
            "materia": row[2] or "",
            "anio": row[3],
            "semestre": row[4] or "",
            "fecha": row[5],
            "nota": row[6],
        })

    # Datos de alumnas
    ws = wb["datos"]
    datos = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[2] is None and row[1] is None:
            continue
        datos.append({
            "nacionalidad": row[0] or "",
            "nombre_religioso": row[1] or "",
            "alumna": row[2] or "",
            "num_documento": row[3] or "",
            "tipo_documento": row[4] or "",
            "fecha_nacimiento": row[5],
        })

    # Profesores
    ws = wb["profesores"]
    profesores = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[1]:
            profesores.append(row[1])

    # Plan de estudios
    ws = wb["Plan de estudios"]
    plan = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[1] is None:
            continue
        codigo = (row[0] or "").strip()
        plan.append({
            "codigo": codigo,
            "materia": row[1],
            "orden": row[2],
            "anio_desc": row[3] or "",
            "ects": row[4],
        })

    wb.close()
    return notas, datos, profesores, plan
