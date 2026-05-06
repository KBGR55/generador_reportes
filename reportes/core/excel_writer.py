"""Escritura segura sobre esquema.xlsx.

- Detecta si el archivo está bloqueado por Excel/LibreOffice.
- Hace un backup con timestamp antes de cada escritura.
- Anexa filas a las hojas correspondientes.
- Asigna un ID único a cada registro nuevo (col `id` al final de la hoja).
  Los registros previos quedan intactos: solo los registros nuevos
  reciben ID. El ID tiene el formato `<prefijo>-<6 dígitos>`, p. ej.
  `NOT-000001`, `ALU-000001`, `PRO-000001`.
"""
import datetime
import os
import shutil

import openpyxl

from reportes.config import EXCEL_PATH, PROJECT_ROOT


BACKUPS_DIR = os.path.join(PROJECT_ROOT, "backups")
ID_HEADER = "id"
ID_PREFIXES = {
    "notas": "NOT",
    "datos": "ALU",
    "profesores": "PRO",
}


class ExcelLockedError(Exception):
    """El archivo esquema.xlsx está abierto en otra aplicación."""


def _lock_file_path():
    folder = os.path.dirname(EXCEL_PATH)
    name = os.path.basename(EXCEL_PATH)
    return os.path.join(folder, f".~lock.{name}#")


def excel_is_locked():
    """True si LibreOffice/Excel tienen el archivo abierto."""
    return os.path.exists(_lock_file_path())


def backup_excel():
    """Copia esquema.xlsx a backups/ con timestamp y devuelve la ruta."""
    os.makedirs(BACKUPS_DIR, exist_ok=True)
    ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    dest = os.path.join(BACKUPS_DIR, f"esquema_{ts}.xlsx")
    shutil.copy2(EXCEL_PATH, dest)
    return dest


def _id_column_index(ws):
    """Devuelve el índice (1-based) de la columna `id` o None si no existe."""
    if ws.max_row < 1:
        return None
    for col_idx, cell in enumerate(ws[1], start=1):
        if cell.value and str(cell.value).strip().lower() == ID_HEADER:
            return col_idx
    return None


def _ensure_id_column(ws):
    """Garantiza que exista una columna `id` y devuelve su índice."""
    idx = _id_column_index(ws)
    if idx is None:
        idx = ws.max_column + 1
        ws.cell(row=1, column=idx, value=ID_HEADER)
    return idx


def _next_id(ws, prefix, id_col_idx):
    """Próximo ID del estilo NOT-000042 (escanea la columna)."""
    max_n = 0
    for row in range(2, ws.max_row + 1):
        val = ws.cell(row=row, column=id_col_idx).value
        if val and str(val).startswith(prefix + "-"):
            try:
                n = int(str(val).split("-", 1)[1])
                max_n = max(max_n, n)
            except ValueError:
                continue
    return f"{prefix}-{max_n + 1:06d}"


def _append_row(sheet_name, row_values):
    """Agrega una fila a la hoja, con ID auto-generado si corresponde.

    Devuelve el ID asignado (o None si la hoja no usa IDs).
    """
    if excel_is_locked():
        raise ExcelLockedError(
            "El archivo esquema.xlsx está abierto en Excel/LibreOffice. "
            "Cerralo antes de guardar."
        )
    backup_excel()

    wb = openpyxl.load_workbook(EXCEL_PATH)
    try:
        ws = wb[sheet_name]
        new_id = None
        if sheet_name in ID_PREFIXES:
            id_col = _ensure_id_column(ws)
            new_id = _next_id(ws, ID_PREFIXES[sheet_name], id_col)

        ws.append(row_values)
        new_row = ws.max_row

        if new_id is not None:
            ws.cell(row=new_row, column=id_col, value=new_id)

        wb.save(EXCEL_PATH)
    finally:
        wb.close()

    return new_id


def add_nota(profesor, alumna, materia, anio, semestre, fecha, nota):
    """Anexa una fila a la hoja `notas`. Devuelve el ID asignado."""
    return _append_row(
        "notas", [profesor, alumna, materia, anio, semestre, fecha, nota])


def add_alumna(nacionalidad, nombre_religioso, alumna,
               num_documento, tipo_documento, fecha_nacimiento):
    """Anexa una fila a la hoja `datos`. Devuelve el ID asignado."""
    return _append_row(
        "datos",
        [nacionalidad, nombre_religioso, alumna,
         num_documento, tipo_documento, fecha_nacimiento])


def add_profesor(nombre):
    """Anexa una fila a la hoja `profesores`. Devuelve el ID asignado."""
    return _append_row("profesores", [None, nombre])
